import pandas as pd
import os
import json
from pathlib import Path
import argparse
from typing import List, Dict, Any, Set, Optional, Tuple


def load_captions_from_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """
    Load captions data from an XLSX file and convert each row to a dictionary.
    
    Args:
        file_path: Path to the XLSX file
        
    Returns:
        List of dictionaries, where each dictionary represents a row with standardized keys
    """
    # Read the Excel file
    df = pd.read_excel(file_path, engine='openpyxl')
    
    # Log column names for debugging
    print(f"\nColumns in {Path(file_path).name}:")
    for i, col in enumerate(df.columns):
        print(f"{i}: {col}")
    
    # Column mapping dictionary for flexibility
    column_mappings = {
        'content_id': ['content_id'],
        'stock_url': ['stock_url'],
        'summary_caption': ['summary', 'what is the video about'],
        'subject_background_caption': ['subject background', 'who or what is in the image'],
        'subject_motion_caption': ['subject action', 'what are they doing'],
        'camera_caption': ['camera', 'how is it captured']
    }
    
    result = []
    
    for _, row in df.iterrows():
        caption_dict = {}
        
        for target_key, possible_matches in column_mappings.items():
            for col in df.columns:
                col_lower = col.lower()
                if any(match in col_lower for match in possible_matches):
                    caption_dict[target_key] = row[col]
                    break
        
        result.append(caption_dict)
    
    return result


def extract_filename(path: str) -> str:
    """Extract the filename from a path or URL."""
    return path.split('/')[-1]


def get_caption_dict_by_filename(captions_data: List[Dict[str, Any]], filename: str) -> Optional[Dict[str, Any]]:
    """
    Find the caption dictionary for a specific filename.
    
    Args:
        captions_data: List of caption dictionaries
        filename: Video filename to look for
        
    Returns:
        Caption dictionary if found, None otherwise
    """
    for caption in captions_data:
        if 'stock_url' in caption and extract_filename(caption['stock_url']) == filename:
            return caption
    return None


def identify_invalid_videos(video_data_dict: Dict[str, Any], label_collections: List[str]) -> Set[str]:
    """
    Identify videos that have shot_transition labels, which are considered invalid.
    
    Args:
        video_data_dict: Dictionary of video data
        label_collections: List of label collections to check
        
    Returns:
        Set of filenames for invalid videos
    """
    # Import only when needed
    from label import Label, extract_labels_dict
    
    # Get all valid labels
    video_data_list = list(video_data_dict.values())
    labels = Label.load_all_labels()
    all_valid_labels_dict = {}
    
    for label_collection in label_collections:
        labels_dict = extract_labels_dict(
            getattr(labels, label_collection), label_collection
        )
        invalid_labels = []
        for label_name, label in labels_dict.items():
            # call verify()
            try:
                label.verify(video_data_list)
            except Exception as e:
                print(f"Skipping {label_name}: {e}")
                invalid_labels.append(label_name)
                continue
        if len(invalid_labels) > 0:
            print(
                f"Skipping {len(invalid_labels)} invalid labels for {label_collection}: {invalid_labels}"
            )
            for label_name in invalid_labels:
                labels_dict.pop(label_name)
        all_valid_labels_dict.update(labels_dict)
    
    # Find shot_transition labels across all label collections
    shot_transition_labels = []
    for label_name in all_valid_labels_dict:
        if "shot_transition" in label_name:
            shot_transition_labels.append(all_valid_labels_dict[label_name])
            print(f"Found shot_transition label: {label_name}")
    
    if not shot_transition_labels:
        print("Warning: No shot_transition labels found!")
        return set()
    
    # Check each video for shot_transition label
    invalid_videos = set()
    for video_name, video_data in video_data_dict.items():
        for label in shot_transition_labels:
            if video_data in label.pos(video_data_list):
                invalid_videos.add(video_name)
                break
    
    print(f"Found {len(invalid_videos)} invalid videos with shot_transition labels")
    return invalid_videos


def analyze_data_overlap(video_data_dict: Dict[str, Any], captions_data: List[Dict[str, Any]], invalid_videos: Set[str]) -> Tuple[List[str], Set[str], Set[str], Set[str], Set[str]]:
    """
    Analyze the overlap between video data and captions data and return a consistently sorted list.
    
    Args:
        video_data_dict: Dictionary of video data
        captions_data: List of caption dictionaries
        invalid_videos: Set of invalid video filenames
        
    Returns:
        Tuple of (sorted_overlap_list, missing_in_captions, missing_in_videos, overlap_invalid, nonoverlap_invalid)
    """
    # Extract filenames from both datasets
    video_names = {extract_filename(name) for name in video_data_dict.keys()}
    stock_names = {extract_filename(caption['stock_url']) for caption in captions_data if 'stock_url' in caption}
    
    # Find overlaps and differences
    overlap = video_names.intersection(stock_names)
    missing_in_captions = video_names - overlap
    missing_in_videos = stock_names - overlap
    
    # Identify invalid videos in overlap and non-overlap sets
    overlap_invalid = overlap.intersection(invalid_videos)
    nonoverlap_invalid = missing_in_captions.intersection(invalid_videos)
    
    # Remove invalid videos from overlap and missing_in_captions
    overlap = overlap - overlap_invalid
    missing_in_captions = missing_in_captions - nonoverlap_invalid
    
    # Sort the overlap for consistent ordering - THIS IS THE MASTER SORT THAT WILL BE USED EVERYWHERE
    sorted_overlap_list = sorted(overlap)
    
    # Print analysis
    print(f"\nOverlap analysis:")
    print(f"Videos in both datasets: {len(sorted_overlap_list)}")
    print(f"Videos in our data but not in Adobe data: {len(missing_in_captions)}")
    print(f"Videos in Adobe data but not in our data: {len(missing_in_videos)}")
    print(f"Invalid videos in overlap: {len(overlap_invalid)}")
    print(f"Invalid videos in non-overlap: {len(nonoverlap_invalid)}")
    
    # Print some example filenames for verification
    if sorted_overlap_list:
        print(f"\nSample overlapping videos (first 5):")
        for i, filename in enumerate(sorted_overlap_list[:5]):
            print(f"  {i+1}. {filename}")
    
    return sorted_overlap_list, missing_in_captions, missing_in_videos, overlap_invalid, nonoverlap_invalid


def save_overlapping_videos_json(sorted_overlap: List[str], save_dir: str, dataset_base_url: str, num_splits: int = 10) -> None:
    """
    Save the overlapping video URLs to JSON files.
    
    Args:
        sorted_overlap: List of overlapping video filenames (already sorted)
        save_dir: Directory to save the JSON files
        dataset_base_url: Base URL for the dataset
        num_splits: Number of splits to create
    """
    # Create directory if it doesn't exist
    os.makedirs(save_dir, exist_ok=True)
    
    # Prepare full URLs using the already sorted list
    full_urls = [f"{dataset_base_url}/{video}" for video in sorted_overlap]
    
    # Save all videos to a single file
    all_videos_path = os.path.join(save_dir, f"overlap_all_{len(full_urls)}.json")
    with open(all_videos_path, 'w') as f:
        json.dump(full_urls, f, indent=4)
    print(f"Saved all overlap videos to {all_videos_path}")
    
    # Save videos in splits
    videos_count = len(full_urls)
    for i in range(num_splits):
        start = i * videos_count // num_splits
        end = (i + 1) * videos_count // num_splits
        
        split_path = os.path.join(save_dir, f"overlap_{start}_to_{end}.json")
        with open(split_path, 'w') as f:
            json.dump(full_urls[start:end], f, indent=4)
        print(f"Saved split {i+1}/{num_splits} to {split_path}")


def save_invalid_videos_json(invalid_videos: Set[str], save_dir: str, dataset_base_url: str, prefix: str) -> None:
    """
    Save the invalid video URLs to JSON files.
    
    Args:
        invalid_videos: Set of invalid video filenames
        save_dir: Directory to save the JSON files
        dataset_base_url: Base URL for the dataset
        prefix: Prefix for the output filename (overlap_invalid or nonoverlap_invalid)
    """
    # Create directory if it doesn't exist
    os.makedirs(save_dir, exist_ok=True)
    
    # Prepare full URLs
    full_urls = [f"{dataset_base_url}/{video}" for video in invalid_videos]
    
    # Save to a single file
    invalid_path = os.path.join(save_dir, f"{prefix}.json")
    with open(invalid_path, 'w') as f:
        json.dump(full_urls, f, indent=4)
    print(f"Saved {len(full_urls)} invalid videos to {invalid_path}")


def format_excel_file(df: pd.DataFrame, output_path: str) -> None:
    """
    Apply formatting to an Excel file to make it more readable.
    
    Args:
        df: DataFrame that was saved to the Excel file
        output_path: Path to the Excel file to format
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Load the workbook
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    
    # Set column widths based on content
    for i, column in enumerate(df.columns):
        column_letter = get_column_letter(i + 1)
        
        # Calculate max width needed (between header and max content)
        max_length = max(
            df[column].astype(str).apply(len).max() if len(df) > 0 else 0,
            len(str(column))
        )
        
        # Set column width (with some padding, max 100 characters)
        adjusted_width = min(max_length + 3, 100)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Format the header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Enable text wrapping for data cells and auto-height for rows
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    for row in range(2, len(df) + 2):
        # Set a reasonable fixed height for rows with content
        worksheet.row_dimensions[row].height = 60  # Default height for content rows
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = wrap_alignment
            
            # Add a light border to all cells
            cell.border = Border(
                left=Side(style='thin', color='DDDDDD'), 
                right=Side(style='thin', color='DDDDDD'), 
                top=Side(style='thin', color='DDDDDD'), 
                bottom=Side(style='thin', color='DDDDDD')
            )
    
    # Freeze the header row
    worksheet.freeze_panes = 'A2'
    
    # Save the formatted workbook
    workbook.save(output_path)


def create_overlap_excel(captions_data: List[Dict[str, Any]], 
                         sorted_overlap: List[str], 
                         xlsx_files: List[str], 
                         output_path: str) -> None:
    """
    Create a new Excel file containing only the overlapping videos with the same structure as the original files.
    The Excel file is formatted to be easily readable, with proper column widths and text wrapping.
    
    Args:
        captions_data: List of all caption dictionaries
        sorted_overlap: List of overlapping video filenames (already sorted)
        xlsx_files: List of original Excel files (to extract column structure)
        output_path: Path to save the new Excel file
    """
    # Read the first Excel file to get its structure
    template_df = pd.read_excel(xlsx_files[0], engine='openpyxl')
    columns = template_df.columns
    
    # Create a new DataFrame with the same columns
    overlap_df = pd.DataFrame(columns=columns)
    
    # Collect corresponding rows from captions_data - using the pre-sorted list
    for filename in sorted_overlap:
        caption_dict = get_caption_dict_by_filename(captions_data, filename)
        if caption_dict:
            # Create a new row with column mapping reversed
            row_data = {}
            
            # Direct mappings for standard columns
            if 'content_id' in caption_dict and 'content_id' in columns:
                row_data['content_id'] = caption_dict['content_id']
            if 'stock_url' in caption_dict and 'stock_url' in columns:
                row_data['stock_url'] = caption_dict['stock_url']
            
            # Map the caption fields to the original column names
            for col in columns:
                col_lower = col.lower()
                if ('summary' in col_lower or 'what is the video about' in col_lower) and 'summary_caption' in caption_dict:
                    row_data[col] = caption_dict['summary_caption']
                elif (('subject' in col_lower and 'background' in col_lower) or 'who or what is in the image' in col_lower) and 'subject_background_caption' in caption_dict:
                    row_data[col] = caption_dict['subject_background_caption']
                elif (('subject' in col_lower and 'action' in col_lower) or 'what are they doing' in col_lower) and 'subject_motion_caption' in caption_dict:
                    row_data[col] = caption_dict['subject_motion_caption']
                elif ('camera' in col_lower or 'how is it captured' in col_lower) and 'camera_caption' in caption_dict:
                    row_data[col] = caption_dict['camera_caption']
            
            # Add the row to the DataFrame
            overlap_df = pd.concat([overlap_df, pd.DataFrame([row_data])], ignore_index=True)
    
    # Create a directory for the output file if it doesn't exist
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # First save the DataFrame to Excel
    overlap_df.to_excel(output_path, index=False, engine='openpyxl')
    
    # Apply formatting to the Excel file
    format_excel_file(overlap_df, output_path)
    
    print(f"Created formatted overlap Excel file with {len(overlap_df)} rows at {output_path}")


def create_overlap_excel_splits(captions_data: List[Dict[str, Any]], 
                                sorted_overlap: List[str], 
                                xlsx_files: List[str], 
                                base_output_path: str,
                                num_splits: int = 10) -> None:
    """
    Create multiple Excel files containing the overlapping videos split into chunks.
    Each file maintains the same structure as the original files and is formatted for readability.
    
    Args:
        captions_data: List of all caption dictionaries
        sorted_overlap: List of overlapping video filenames (already sorted)
        xlsx_files: List of original Excel files (to extract column structure)
        base_output_path: Base path for the Excel files (will add split suffixes)
        num_splits: Number of splits to create
    """
    # Read the first Excel file to get its structure
    template_df = pd.read_excel(xlsx_files[0], engine='openpyxl')
    columns = template_df.columns
    
    # Use the already sorted list
    total_videos = len(sorted_overlap)
    
    # Ensure base_output_path directory exists
    base_dir = os.path.dirname(base_output_path)
    os.makedirs(base_dir, exist_ok=True)
    
    # Calculate videos per split
    for i in range(num_splits):
        start = i * total_videos // num_splits
        end = (i + 1) * total_videos // num_splits
        
        # Get the filenames for this split
        split_filenames = sorted_overlap[start:end]
        
        # Create output path for this split - ensure it has .xlsx extension
        split_output_path = os.path.join(base_dir, f"overlap_split_{start}_to_{end}.xlsx")
        
        # Create a new DataFrame with the same columns
        split_df = pd.DataFrame(columns=columns)
        
        # Collect corresponding rows for this split
        for filename in split_filenames:
            caption_dict = get_caption_dict_by_filename(captions_data, filename)
            if caption_dict:
                # Create a new row with column mapping reversed
                row_data = {}
                
                # Direct mappings for standard columns
                if 'content_id' in caption_dict and 'content_id' in columns:
                    row_data['content_id'] = caption_dict['content_id']
                if 'stock_url' in caption_dict and 'stock_url' in columns:
                    row_data['stock_url'] = caption_dict['stock_url']
                
                # Map the caption fields to the original column names
                for col in columns:
                    col_lower = col.lower()
                    if ('summary' in col_lower or 'what is the video about' in col_lower) and 'summary_caption' in caption_dict:
                        row_data[col] = caption_dict['summary_caption']
                    elif (('subject' in col_lower and 'background' in col_lower) or 'who or what is in the image' in col_lower) and 'subject_background_caption' in caption_dict:
                        row_data[col] = caption_dict['subject_background_caption']
                    elif (('subject' in col_lower and 'action' in col_lower) or 'what are they doing' in col_lower) and 'subject_motion_caption' in caption_dict:
                        row_data[col] = caption_dict['subject_motion_caption']
                    elif ('camera' in col_lower or 'how is it captured' in col_lower) and 'camera_caption' in caption_dict:
                        row_data[col] = caption_dict['camera_caption']
                
                # Add the row to the DataFrame
                split_df = pd.concat([split_df, pd.DataFrame([row_data])], ignore_index=True)
        
        if len(split_df) > 0:
            # First save the DataFrame to Excel
            split_df.to_excel(split_output_path, index=False, engine='openpyxl')
            
            # Then open the file with openpyxl and apply formatting
            format_excel_file(split_df, split_output_path)
            
            print(f"Created split {i+1}/{num_splits} Excel file with {len(split_df)} rows at {split_output_path}")
        else:
            print(f"Skip creating split {i+1}/{num_splits} Excel file as no data found")


def save_nonoverlapping_videos_json(missing_in_videos: Set[str], save_dir: str, dataset_base_url: str) -> None:
    """
    Save the non-overlapping video URLs to a JSON file.
    
    Args:
        missing_in_videos: Set of video filenames that are in our data but not in Adobe data
        save_dir: Directory to save the JSON file
        dataset_base_url: Base URL for the dataset
    """
    # Create directory if it doesn't exist
    os.makedirs(save_dir, exist_ok=True)
    
    # Prepare full URLs
    full_urls = [f"{dataset_base_url}/{video}" for video in missing_in_videos]
    
    # Save to a single file
    nonoverlap_path = os.path.join(save_dir, f"nonoverlap_all_{len(full_urls)}.json")
    with open(nonoverlap_path, 'w') as f:
        json.dump(full_urls, f, indent=4)
    print(f"Saved {len(full_urls)} non-overlapping videos to {nonoverlap_path}")


def parse_args():
    """Parse command line arguments for load_xlsx script"""
    parser = argparse.ArgumentParser(description="Load and process Excel caption data")
    parser.add_argument("--video_data", type=str, default="video_data/20250406_setup_and_motion/videos.json", 
                       help="Path to the video data file")
    parser.add_argument("--label_collections", nargs="+", type=str, default=["cam_motion", "cam_setup", "lighting_setup"], 
                       help="List of label collections to load from the video data")
    return parser.parse_args()

def main():
    # Import only when needed
    from process_json import json_to_video_data
    
    # Parse command line arguments
    args = parse_args()
    
    # Configuration
    xlsx_files = ['adobe_2_19.xlsx', 'adobe_2_17.xlsx']
    dataset_base_url = "https://huggingface.co/datasets/zhiqiulin/video_captioning/resolve/main"
    
    # Load caption data from all xlsx files
    print("Loading caption data from Excel files...")
    captions_data = []
    for file_path in xlsx_files:
        captions_data.extend(load_captions_from_xlsx(file_path))
    
    print(f"Total caption records loaded: {len(captions_data)}")
    
    # Load video data
    print("Loading video data...")
    video_data_dict = json_to_video_data(args.video_data, label_collections=args.label_collections)
    
    # Identify invalid videos with shot_transition labels
    print("Identifying invalid videos with shot_transition labels...")
    invalid_videos = identify_invalid_videos(video_data_dict, args.label_collections)
    
    # Analyze data overlap and get a SINGLE sorted list that will be used everywhere
    sorted_overlap, missing_in_captions, missing_in_videos, overlap_invalid, nonoverlap_invalid = analyze_data_overlap(video_data_dict, captions_data, invalid_videos)
    
    # Extract video data name from path
    video_data_name = Path(args.video_data).parent.name
    print(f"Processing video data from: {video_data_name}")
    
    # Define save directories
    json_dir = f"caption/video_urls/{video_data_name}/"
    excel_dir = f"caption/excel_export/{video_data_name}/"
    os.makedirs(excel_dir, exist_ok=True)
    
    # Save overlapping videos to JSON files - using the SAME sorted list
    save_overlapping_videos_json(sorted_overlap, json_dir, dataset_base_url)
    
    # Save non-overlapping videos to JSON file
    save_nonoverlapping_videos_json(missing_in_captions, json_dir, dataset_base_url)
    
    # Save invalid videos to JSON files
    save_invalid_videos_json(overlap_invalid, json_dir, dataset_base_url, "overlap_invalid")
    save_invalid_videos_json(nonoverlap_invalid, json_dir, dataset_base_url, "nonoverlap_invalid")
    
    # Save overlapping videos to main Excel file - using the SAME sorted list
    excel_output_path = os.path.join(excel_dir, "overlap_all.xlsx")
    create_overlap_excel(captions_data, sorted_overlap, xlsx_files, excel_output_path)
    
    # Save split Excel files - using the SAME sorted list
    split_base_path = os.path.join(excel_dir, "overlap")
    create_overlap_excel_splits(captions_data, sorted_overlap, xlsx_files, split_base_path, 10)
    
    print(f"Processing complete. Files saved to:")
    print(f"  - JSON files: {json_dir}")
    print(f"  - Excel files: {excel_dir}")


if __name__ == "__main__":
    main()