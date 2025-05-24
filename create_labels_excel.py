import os
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

def get_all_json_files():
    json_files = []
    labels_dir = Path('labels')
    
    # Walk through all subdirectories
    for root, dirs, files in os.walk(labels_dir):
        for file in files:
            if file.endswith('.json'):
                # Get relative path from labels directory
                rel_path = Path(root).relative_to(labels_dir)
                json_files.append((rel_path, file))
    
    return json_files

def create_excel():
    json_files = get_all_json_files()
    
    # Create a list to store all data
    data = []
    data_with_def = []
    
    # Process each JSON file
    for rel_path, file in json_files:
        row = {}
        row_with_def = {}
        # Add each directory level as a column
        for i, part in enumerate(rel_path.parts):
            row[f'Level_{i+1}'] = part
            row_with_def[f'Level_{i+1}'] = part
        # Add filename without extension as Label
        row['Label'] = file[:-5]  # Remove .json extension
        row_with_def['Label'] = file[:-5]
        
        # Read the JSON file to get the definition
        json_path = Path('labels') / rel_path / file
        with open(json_path, 'r') as f:
            json_data = json.load(f)
            # Get the first definition from def_prompt
            definition = json_data.get('def_prompt', [''])[0] if json_data.get('def_prompt') else ''
            row_with_def['Definition'] = definition
        
        data.append(row)
        data_with_def.append(row_with_def)
    
    # Create DataFrames
    df = pd.DataFrame(data)
    df_with_def = pd.DataFrame(data_with_def)
    
    # Define the order of main categories
    category_order = ['cam_motion', 'cam_setup', 'lighting_setup']
    
    # Sort the DataFrames
    df['sort_key'] = df['Level_1'].map(lambda x: category_order.index(x) if x in category_order else len(category_order))
    df = df.sort_values(['sort_key', 'Level_1', 'Level_2', 'Level_3', 'Label'])
    df = df.drop('sort_key', axis=1)
    
    df_with_def['sort_key'] = df_with_def['Level_1'].map(lambda x: category_order.index(x) if x in category_order else len(category_order))
    df_with_def = df_with_def.sort_values(['sort_key', 'Level_1', 'Level_2', 'Level_3', 'Label'])
    df_with_def = df_with_def.drop('sort_key', axis=1)
    
    # Add Number column at the end
    df['Number'] = ''
    
    # Reorder columns
    cols = [col for col in df.columns if col not in ['Label', 'Number']] + ['Label', 'Number']
    df = df[cols]
    
    cols_with_def = [col for col in df_with_def.columns if col not in ['Label', 'Definition']] + ['Label', 'Definition']
    df_with_def = df_with_def[cols_with_def]
    
    # Create Excel writer
    with pd.ExcelWriter('labels_structure.xlsx', engine='openpyxl') as writer:
        # Write both DataFrames to different sheets
        df.to_excel(writer, index=False, sheet_name='Labels Structure')
        df_with_def.to_excel(writer, index=False, sheet_name='Labels with Definitions')
        
        # Get the workbook and worksheets
        workbook = writer.book
        worksheet = writer.sheets['Labels Structure']
        worksheet_def = writer.sheets['Labels with Definitions']
        
        # Define styles
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format both worksheets
        for ws in [worksheet, worksheet_def]:
            # Format the header
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Format data cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
            
            # Group and merge cells by main category
            current_category = None
            start_row = 2
            category_rows = {}
            
            # First pass: collect row ranges for each category
            for row in range(2, ws.max_row + 1):
                category = ws.cell(row=row, column=1).value  # Level_1 column
                
                if category != current_category:
                    if current_category is not None:
                        category_rows[current_category] = (start_row, row - 1)
                    current_category = category
                    start_row = row
            
            # Add the last category
            if current_category is not None:
                category_rows[current_category] = (start_row, ws.max_row)
            
            # Second pass: merge cells for each category
            for category, (start, end) in category_rows.items():
                # Merge cells in the first column (Level_1)
                ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
                merged_cell = ws.cell(row=start, column=1)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                merged_cell.fill = subheader_fill
                merged_cell.font = Font(bold=True)
                
                # Also merge cells in Level_2 if they're the same
                current_level2 = None
                level2_start = start
                for row in range(start, end + 1):
                    level2 = ws.cell(row=row, column=2).value  # Level_2 column
                    if level2 != current_level2:
                        if current_level2 is not None and row - level2_start > 1:
                            ws.merge_cells(start_row=level2_start, start_column=2, end_row=row-1, end_column=2)
                            merged_cell = ws.cell(row=level2_start, column=2)
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                            merged_cell.fill = subheader_fill
                            merged_cell.font = Font(bold=True)
                            
                            # Merge Level_3 cells within this Level_2 group
                            current_level3 = None
                            level3_start = level2_start
                            for row3 in range(level2_start, row):
                                level3 = ws.cell(row=row3, column=3).value
                                if level3 != current_level3:
                                    if current_level3 is not None and row3 - level3_start > 1:
                                        ws.merge_cells(start_row=level3_start, start_column=3, end_row=row3-1, end_column=3)
                                        merged_cell = ws.cell(row=level3_start, column=3)
                                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                                        merged_cell.fill = subheader_fill
                                        merged_cell.font = Font(bold=True)
                                    current_level3 = level3
                                    level3_start = row3
                            
                            # Merge the last group of Level_3 cells in this Level_2 group
                            if row - level3_start > 1:
                                ws.merge_cells(start_row=level3_start, start_column=3, end_row=row-1, end_column=3)
                                merged_cell = ws.cell(row=level3_start, column=3)
                                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                                merged_cell.fill = subheader_fill
                                merged_cell.font = Font(bold=True)
                        
                        current_level2 = level2
                        level2_start = row
                
                # Merge the last group of Level_2 cells
                if end - level2_start > 0:
                    ws.merge_cells(start_row=level2_start, start_column=2, end_row=end, end_column=2)
                    merged_cell = ws.cell(row=level2_start, column=2)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                    merged_cell.fill = subheader_fill
                    merged_cell.font = Font(bold=True)
                    
                    # Merge Level_3 cells within the last Level_2 group
                    current_level3 = None
                    level3_start = level2_start
                    for row3 in range(level2_start, end + 1):
                        level3 = ws.cell(row=row3, column=3).value
                        if level3 != current_level3:
                            if current_level3 is not None and row3 - level3_start > 1:
                                ws.merge_cells(start_row=level3_start, start_column=3, end_row=row3-1, end_column=3)
                                merged_cell = ws.cell(row=level3_start, column=3)
                                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                                merged_cell.fill = subheader_fill
                                merged_cell.font = Font(bold=True)
                            current_level3 = level3
                            level3_start = row3
                    
                    # Merge the last group of Level_3 cells
                    if end - level3_start > 0:
                        ws.merge_cells(start_row=level3_start, start_column=3, end_row=end, end_column=3)
                        merged_cell = ws.cell(row=level3_start, column=3)
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                        merged_cell.fill = subheader_fill
                        merged_cell.font = Font(bold=True)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Freeze the header row
            ws.freeze_panes = 'A2'
            
            # Special formatting for the Definition column in the second sheet
            if ws == worksheet_def:
                # Set a larger width for the Definition column
                ws.column_dimensions['E'].width = 60  # Adjust the width for definitions
                
                # Format the Definition column cells
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='center')

if __name__ == '__main__':
    create_excel() 